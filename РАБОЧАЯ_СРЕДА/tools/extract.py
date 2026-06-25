"""
extract.py — извлечение содержимого исходников в редактируемый Markdown.

Поддерживает .pdf (текст + таблицы), .docx (абзацы, заголовки, таблицы),
.xlsx (листы → markdown-таблицы). Пишет .md рядом или в указанный путь.

Примеры:
    python tools/extract.py "../материалы/шаблоны_конкурса/Application_Form.docx"
    python tools/extract.py "../материалы/шаблоны_конкурса/Budget_template.xlsx" -o source_extracted/Budget.md
    python tools/extract.py "../материалы/техника/04_Технические_выжимки_и_схемы.pdf"
    python tools/extract.py --all          # извлечь весь набор шаблонов конкурса
"""
from __future__ import annotations
import argparse
import sys
from pathlib import Path

import common


# ── PDF ────────────────────────────────────────────────────────────────────
def pdf_to_md(path: Path) -> str:
    import fitz  # pymupdf
    import pdfplumber

    out = [f"# {path.stem}\n\n> Извлечено из `{path.name}` ({path.suffix}).\n"]
    doc = fitz.open(path)
    # Текст постранично через pymupdf
    page_texts = [page.get_text("text") for page in doc]
    doc.close()
    # Таблицы через pdfplumber (надёжнее для табличных шаблонов)
    tables_by_page: dict[int, list] = {}
    try:
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages):
                tbls = page.extract_tables()
                if tbls:
                    tables_by_page[i] = tbls
    except Exception as e:  # noqa: BLE001
        out.append(f"\n> ⚠️ pdfplumber не смог разобрать таблицы: {e}\n")

    for i, text in enumerate(page_texts):
        out.append(f"\n## Стр. {i + 1}\n")
        if text.strip():
            out.append(text.rstrip())
        for t_idx, table in enumerate(tables_by_page.get(i, [])):
            out.append(f"\n*Таблица {i + 1}.{t_idx + 1}:*\n")
            out.append(_rows_to_md(table))
    return "\n".join(out) + "\n"


# ── DOCX ───────────────────────────────────────────────────────────────────
def docx_to_md(path: Path) -> str:
    from docx import Document
    from docx.document import Document as _Doc
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    doc = Document(str(path))
    out = [f"# {path.stem}\n\n> Извлечено из `{path.name}`.\n"]

    def iter_block_items(parent):
        """Итерируем абзацы и таблицы в порядке документа."""
        from docx.oxml.table import CT_Tbl
        from docx.oxml.text.paragraph import CT_P
        body = parent.element.body
        for child in body.iterchildren():
            if isinstance(child, CT_P):
                yield Paragraph(child, parent)
            elif isinstance(child, CT_Tbl):
                yield Table(child, parent)

    for block in iter_block_items(doc):
        if isinstance(block, Paragraph):
            txt = block.text.rstrip()
            if not txt:
                continue
            style = (block.style.name or "").lower()
            if style.startswith("heading 1") or style == "title":
                out.append(f"\n## {txt}")
            elif style.startswith("heading 2"):
                out.append(f"\n### {txt}")
            elif style.startswith("heading 3"):
                out.append(f"\n#### {txt}")
            elif style.startswith("list"):
                out.append(f"- {txt}")
            else:
                out.append(txt)
        elif isinstance(block, Table):
            rows = [[c.text.strip().replace("\n", " ") for c in r.cells] for r in block.rows]
            out.append("")
            out.append(_rows_to_md(rows))
    return "\n".join(out) + "\n"


# ── XLSX ───────────────────────────────────────────────────────────────────
def xlsx_to_md(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), data_only=True)
    out = [f"# {path.stem}\n\n> Извлечено из `{path.name}` (значения формул вычислены).\n"]
    for ws in wb.worksheets:
        out.append(f"\n## Лист: {ws.title}\n")
        rows = []
        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            cells = ["" if v is None else str(v).replace("\n", " ").strip() for v in row]
            if any(c for c in cells):
                rows.append(cells)
        if not rows:
            out.append("*(пустой лист)*")
            continue
        # выровнять ширину строк
        width = max(len(r) for r in rows)
        rows = [r + [""] * (width - len(r)) for r in rows]
        out.append(_rows_to_md(rows))
    return "\n".join(out) + "\n"


# ── helpers ────────────────────────────────────────────────────────────────
def _rows_to_md(rows: list[list]) -> str:
    rows = [["" if c is None else str(c).replace("|", "\\|").replace("\n", " ").strip()
             for c in r] for r in rows if r is not None]
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    header = rows[0]
    body = rows[1:] if len(rows) > 1 else []
    lines = ["| " + " | ".join(header) + " |",
             "|" + "|".join(["---"] * width) + "|"]
    for r in body:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines)


DISPATCH = {".pdf": pdf_to_md, ".docx": docx_to_md, ".xlsx": xlsx_to_md}


def extract_one(src: Path, dst: Path | None) -> Path:
    fn = DISPATCH.get(src.suffix.lower())
    if not fn:
        raise SystemExit(f"Не поддерживается: {src.suffix} ({src.name})")
    md = fn(src)
    if dst is None:
        dst = common.SOURCE_EXTRACTED / (src.stem + ".md")
    dst.parent.mkdir(parents=True, exist_ok=True)
    dst.write_text(md, encoding="utf-8")
    print(f"✓ {src.name}  →  {dst.relative_to(common.ENV_ROOT) if common.ENV_ROOT in dst.parents else dst}")
    return dst


def main():
    ap = argparse.ArgumentParser(description="Извлечь PDF/DOCX/XLSX в Markdown")
    ap.add_argument("source", nargs="?", help="путь к .pdf/.docx/.xlsx")
    ap.add_argument("-o", "--out", help="путь к .md (по умолчанию source_extracted/<имя>.md)")
    ap.add_argument("--all", action="store_true",
                    help="извлечь все шаблоны конкурса + ключевые PDF")
    args = ap.parse_args()

    if args.all:
        targets = sorted(
            [p for p in common.TEMPLATES.glob("*") if p.suffix.lower() in DISPATCH]
            + [common.MATERIALS / "техника" / "04_Технические_выжимки_и_схемы.pdf"]
        )
        for t in targets:
            if t.exists():
                try:
                    extract_one(t, None)
                except Exception as e:  # noqa: BLE001
                    print(f"✗ {t.name}: {e}", file=sys.stderr)
        return

    if not args.source:
        ap.error("укажите файл или --all")
    extract_one(Path(args.source).resolve(), Path(args.out).resolve() if args.out else None)


if __name__ == "__main__":
    main()
